#!/usr/bin/env python

import math
from openpyxl import Workbook,load_workbook 

f = open("./path.txt")
wb = load_workbook(f.read()[:-1] + '/lyr.xlsx')
f.close()
ws = wb.active

# prompt for either: spending, calculate plat needed for spending target, remaining gold per kill

# Outputs the kills and time remaining until plat goal is reached
def remainingcalc():
    # prompt for current plat
    user = int(raw_input("current plat = "))
    
    # sum total
    total = 0
    for cell in ws['C']:
        if cell.value != None:
            total += cell.value
        else:
            break
    print(str("{:.0f}".format(total)) + "p saved")

    # (FLOOR((10*(C3+total)-c2)/9,1)-c1)/(C4*0.01)
    c8 = math.floor(((10 * (float(ws['E3'].value) + total) - float(ws['E2'].value)) / 9 - user) / (ws['E4'].value * 0.01))
    if c8 < 0:
        print("0 kills")
        print("0 hours 0 minutes 0 seconds")
    else:
        # hours = floor(c8/600)
        # minutes = mod(floor(c8/10),60)
        # seconds = mod(floor(c8*6),60)
        print(str("{:.0f}".format(c8)) + " kills")
        print(str("{:.0f}".format(math.floor(c8/600))) + " hours " + str("{:.0f}".format(math.floor(c8/10) % 60)) + " minutes " + str("{:.0f}".format(math.floor(c8*6) % 60)) + " seconds")

# TARGET
def stcalc():
    # prompt for spending target
    user = raw_input('spending target = ')
    if user != "exit":
        ws['E3'] = int(user)

        # sum all saved income
        total = 0
        for cell in ws['C']:
            if cell.value != None:
                total += cell.value
            else:
                break
        
        # print the plat needed    
        # =FLOOR((10*(H6+total)-H3)/9,1)
        print(str("{:.0f}".format(math.floor((10 * (int(ws['E3'].value) + total) - int(ws['E2'].value)) / 9))) + "p needed")

        # print kills and time remaining
        # remainingcalc()

# REMAINING
def remaining():
    # prompt for gpk
    user = raw_input("gold per kill = ")
    if user != "exit":
        ws['E4'] = float(user)
        stcalc()
        remainingcalc()

# SPENDING
def spending():
    # prompt for plat before spending
    user = raw_input("plat before spending = ")
    if user != "exit":
        # record current plat
        ws['B'+str(len(ws['B'])+1)].value = int(user)
    
        total = 0
        # find empty row
        for cell in ws['C']:
            if cell.value is None:
                # add current plat saved to list
                saved = math.floor((int(user) - int(ws['E2'].value))/10)
                cell.value = saved
                total += saved
                print(str("{:.0f}".format(total)) + "p saved")
                break
            else:
                total += cell.value
        
        # prompt for plat after spending
        pas = int(raw_input("plat after spending = "))

        # set last plat and current plat to plat after spending
        ws['E2'] = pas
        for cell in ws['A']:
            if cell.value is None:
                cell.value = pas
                break



# ask what the user wants to do then again after each action until they exit
print("type exit to quit")
choose = raw_input("spend, target, or remaining? ")
while choose != "exit":
    if choose == "spend":
        spending()
    elif choose == "target":
        stcalc()
    elif choose == "remaining":
        remaining()
    else:
        print("Invalid choice!")
    choose = raw_input("spend, target, or remaining? ")

f = open("path.txt")
wb.save(f.read()[:-1] + "/lyr.xlsx")
f.close()
