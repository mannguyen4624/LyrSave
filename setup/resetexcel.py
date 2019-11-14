#!/usr/bin/env python

import math
from pathlib import Path
from openpyxl import *

f = open("../path.txt")
try:
	wb = load_workbook(f.read()[:-1] + '/lyr.xlsx')
except:
    wb = Workbook()
f.close()
wb.remove(wb.active)
wb.create_sheet()
ws = wb.active
ws['A1'] = '0'
ws['E2'] = '0'
ws['E3'] = '0'

f = open("../path.txt")
wb.save(f.read()[:-1] + "/lyr.xlsx")
f.close()